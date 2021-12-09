import pickle
from tensorflow.keras.models import load_model
import numpy as np
from joblib import load as j_load
from codes.base import eyeing as ey
from scipy import signal
from openpyxl import Workbook
import os


PATH2ROOT = ""


class EyeTrack(object):
    @staticmethod
    def get_pixels(num, testing=False, delete_files=False):
        sbj_dir = PATH2ROOT + f"subjects/{num}/"
        model_boi_dir = sbj_dir + "model-boi"
        scalers_boi_dir = sbj_dir + "scalers-boi.bin"
        model_et_hrz_dir = sbj_dir + "model-et-hrz"
        model_et_vrt_dir = sbj_dir + "model-et-vrt"
        scalers_et_dir = sbj_dir + "scalers-et.bin"
        min_out_ratio = 0.005
        max_out_ratio = 0.995
        y_scale = 1000.0

        if testing:
            sampling_fol = "testing/"
        else:
            sampling_fol = "sampling/"

        sampling_dir = sbj_dir + sampling_fol
        t_load, x1_load, x2_load = ey.load(sampling_dir, ['t', 'x1', 'x2'])

        n_smp = t_load.shape[0]
        print(f"Number of sampling data : {n_smp}")

        # Normalizing Sampling data for 'in_blink_out' model
        x2_chs_inp_boi = x2_load[:, ey.CHOSEN_INPUTS]
        x1_scaler_boi, x2_scaler_boi = j_load(scalers_boi_dir)
        x1_boi = x1_load / x1_scaler_boi
        x2_boi = x2_scaler_boi.transform(x2_chs_inp_boi)

        model_boi = load_model(model_boi_dir)
        y_hat_boi = model_boi.predict([x1_boi, x2_boi]).argmax(1)

        x2_chs_inp_et = x2_load[:, ey.CHOSEN_INPUTS]
        x1_scaler_et, x2_scaler_et = j_load(scalers_et_dir)
        x1_et = x1_load / x1_scaler_et
        x2_et = x2_scaler_et.transform(x2_chs_inp_et)
        x_et = [x1_et, x2_et]

        model_et_hrz = load_model(model_et_hrz_dir)
        model_et_vrt = load_model(model_et_vrt_dir)

        y_hrz_hat = np.expand_dims(model_et_hrz.predict(x_et).reshape((n_smp,)), 1) / y_scale
        y_vrt_hat = np.expand_dims(model_et_vrt.predict(x_et).reshape((n_smp,)), 1) / y_scale
        # y_hrz_hat[y_hrz_hat < min_out_ratio] = min_out_ratio
        # y_vrt_hat[y_vrt_hat < min_out_ratio] = min_out_ratio
        # y_hrz_hat[y_hrz_hat > max_out_ratio] = max_out_ratio
        # y_vrt_hat[y_vrt_hat > max_out_ratio] = max_out_ratio
        y_hat_et = (np.concatenate([y_hrz_hat, y_vrt_hat], 1))

        et_in = y_hat_et.copy()
        # not_in = (y_hat_boi != 2)
        # not_in[0] = False
        # not_in[-1] = False
        # i = 0
        # while i<=n_smp-2:
        #     j = 0
        #     if not_in[i+1]:
        #         while not_in[i+j+1]:
        #             j += 1
        #             if i+j >=n_smp-2:
        #                 break
        #         det = (et_in[i+j+1]-et_in[i])/(j+1)
        #         for (ii, k) in enumerate(range(i+1, i+j+1)):
        #             et_in[k] = et_in[i] + (ii+1) * det
        #     i += j+1

        et_med = et_in.copy()
        et_med[:, 0] = signal.medfilt(et_in[:, 0], 5)
        et_med[:, 1] = signal.medfilt(et_in[:, 1], 5)

        ey.save([et_med], sampling_dir, ['et-flt-in'])

        wb = Workbook()
        ws = wb.active
        ws['A1'] = "# of Sample"
        ws['B1'] = "Time (sec)"
        ws['C1'] = "x (pixel/screen_width)"
        ws['D1'] = "y (pixel/screen_height)"
        i = 0
        while i < n_smp:
            ws[f'A{i+2}'] = i
            ws[f'B{i+2}'] = t_load[i]
            ws[f'C{i+2}'] = et_med[i, 0]
            ws[f'D{i+2}'] = et_med[i, 1]
            i += 1

        wb.save(sampling_dir + "EYE-TRACK.xlsx")

        if delete_files:
            ey.remove(model_boi_dir)
            os.remove(scalers_boi_dir)
            ey.remove(model_et_hrz_dir)
            ey.remove(model_et_vrt_dir)
            os.remove(scalers_et_dir)
            ey.remove(sampling_dir, ['x1', 'x2'])
        else:
            ey.save([y_hat_boi, y_hat_et], sampling_dir, ['y-hat-boi', 'y-hat-et'])


    @staticmethod
    def get_fixations(
            num,
            testing=False,
            t_discard=0.3,
            x_merge=0.1,
            y_merge=0.1,
            vx_thr=2.5,
            vy_thr=2.5
    ):
          # pxr  --> pixel ratio (pixel/screen_width)
          # pxr
          # pxr/sec
          # pxr/sec
        if testing:
            sampling_fol = "testing/"
        else:
            sampling_fol = "sampling/"

        sampling_dir = PATH2ROOT + f"subjects/{num}/" + sampling_fol

        t, et_med = ey.load(sampling_dir, ['t', 'et-flt-in'])

        n_smp = t.shape[0]

        vet_med = et_med.copy()
        vet_med[1:, 0] = (et_med[1:, 0] - et_med[:-1, 0]) / (t[1:] - t[:-1])
        vet_med[1:, 1] = (et_med[1:, 1] - et_med[:-1, 1]) / (t[1:] - t[:-1])
        vet_med[0] = vet_med[1]

        sac = ((vet_med[:, 0] > vx_thr) + (vet_med[:, 0] < -vx_thr)) + (
                    (vet_med[:, 1] > vy_thr) + (vet_med[:, 1] < -vy_thr))
        sac[0] = True
        sac[-1] = True

        i = 0
        fix = []
        while i < n_smp - 1:
            if sac[i]:
                j = 0
                while True:
                    if sac[i + j + 1]:
                        break
                    else:
                        j += 1
                fix.append([i,
                            j,
                            t[i + 1],
                            round(t[i + j + 1] - t[i + 1], 2),
                            round(et_med[i + 1:i + j + 1, 0].mean(), 4),
                            round(et_med[i + 1:i + j + 1, 1].mean(), 4)])
                i += j + 1
        print(fix)

        fix_merge1 = []
        f_d = (x_merge ** 2 + y_merge ** 2) ** 0.5
        n_fix = len(fix)
        i = 0
        not_joined = False
        while i <= n_fix - 2:
            f_new = fix[i]
            j = 1
            while (i + j) <= n_fix:
                fj = fix[i + j]
                fj_d = ((fj[4] - f_new[4]) ** 2 + (fj[5] - f_new[5]) ** 2) ** 0.5
                if fj_d < f_d:
                    f_new = [f_new[0],
                             f_new[1] + fj[1],
                             f_new[2],
                             round(f_new[3] + fj[3], 2),
                             round((f_new[4] * f_new[1] + fj[4] * fj[1]) / (f_new[1] + fj[1]), 4),
                             round((f_new[5] * f_new[1] + fj[5] * fj[1]) / (f_new[1] + fj[1]), 4)]
                    not_joined = False
                else:
                    fix_merge1.append(f_new)
                    not_joined = True
                    break
                j += 1
            i += j

        if not_joined:
            fix_merge1.append(fix[-1])

        print(fix_merge1)

        fix_discard = []
        for f in fix_merge1:
            if f[3] >= t_discard:
                fix_discard.append(f)

        print(fix_discard)

        fix_merge2 = []
        n_fix2 = len(fix_discard)
        i = 0
        not_joined = False
        while i <= n_fix2 - 2:
            f_new = fix_discard[i]
            j = 1
            while (i + j) <= n_fix2:
                fj = fix_discard[i + j]
                fj_d = ((fj[4] - f_new[4]) ** 2 + (fj[5] - f_new[5]) ** 2) ** 0.5
                if fj_d < f_d:
                    f_new = [f_new[0],
                             f_new[1] + fj[1],
                             f_new[2],
                             round(f_new[3] + fj[3], 2),
                             round((f_new[4] * f_new[1] + fj[4] * fj[1]) / (f_new[1] + fj[1]), 4),
                             round((f_new[5] * f_new[1] + fj[5] * fj[1]) / (f_new[1] + fj[1]), 4)]
                    not_joined = False
                else:
                    fix_merge2.append(f_new)
                    not_joined = True
                    break
                j += 1
            i += j

        if not_joined:
            fix_merge2.append(fix_discard[-1])

        print(fix_merge2)

        ey.save([np.array(fix_merge2)], sampling_dir, ['fixations'])

        wb = Workbook()
        ws = wb.active
        ws['A1'] = "# of fixation"
        ws['B1'] = "start of fixation (sample)"
        ws['C1'] = "fixation duration (sample)"
        ws['D1'] = "start of fixation (sec)"
        ws['E1'] = "fixation duration (sec)"
        ws['F1'] = "fixation mean (x)"
        ws['G1'] = "fixation mean (y)"
        
        for (i, f) in enumerate(fix_merge2):
            ws[f'A{i+2}'] = i+1
            ws[f'B{i+2}'] = f[0]
            ws[f'C{i+2}'] = f[1]
            ws[f'D{i+2}'] = f[2]
            ws[f'E{i+2}'] = f[3]
            ws[f'F{i+2}'] = f[4]
            ws[f'G{i+2}'] = f[5]
        
        wb.save(sampling_dir + "FIXATIONS.xlsx")


    @staticmethod
    def get_fix_in_aoi(num, aoi=None):
        sampling_dir = PATH2ROOT + f"subjects/{num}/" + "sampling/"

        if not aoi:
            aoi = [[[0, 0], [0.33, 0.33]], [[0.33, 0], [0.66, 0.33]], [[0.66, 0], [1, 0.33]],
                   [[0, 0.33], [0.33, 0.66]], [[0.33, 0.33], [0.66, 0.66]], [[0.66, 0.33], [1, 0.66]],
                   [[0, 0.66], [0.33, 1]], [[0.33, 0.66], [0.66, 1]], [[0.66, 0.66], [1, 1]]]

        fix = ey.load(sampling_dir, ['fixations'])

        fix_aoi = []
        for f in fix:
            f1 = f.copy()
            xm = f[4]
            ym = f[5]
            for (i, a) in enumerate(aoi):
                if (a[0][0] < xm) and (a[1][0] > xm) and (a[0][1] < ym) and (a[1][1] > ym):
                    f1.append(i + 1)
            fix_aoi.append(f1)

        print(fix_aoi)

        fs_in_aoi = np.zeros((9, 2))

        for f in fix_aoi:
            fs_in_aoi[f[-1] - 1, 0] = fs_in_aoi[f[-1] - 1, 0] + 1
            fs_in_aoi[f[-1] - 1, 1] = fs_in_aoi[f[-1] - 1, 1] + f[3]

        print(fs_in_aoi)

        ey.save([fs_in_aoi], sampling_dir, ['fix-in-aoi'])

        # wb = Workbook()
        # ws = wb.active
        # ws['A1'] = "# of AOI"
        # ws['B1'] = "# of Fixations in AOI"
        # ws['C1'] = "Fixations Time in AOI"
        # for (i, f) in enumerate(fs_in_aoi):
        #     ws[f'A{i+2}'] = i+1
        #     ws[f'B{i+2}'] = f[0]
        #     ws[f'C{i+2}'] = f[1]
        
        wb.save(sampling_dir + "FIXATIONS-IN-AOI.xlsx")