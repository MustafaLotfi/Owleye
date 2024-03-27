"""This file is for creating and training the neural network models for eye movement prediction. Also, this is for creating and training
The in-out model which is for predicting whether the subject is looking inside of the screen or outside of the screen. To understand this
module, you should know about how to build neural network models with keras and tensorflow"""

from tensorflow.keras.layers import (Input, Conv2D, Flatten, MaxPooling2D,
                                     Dense, Dropout, Concatenate)
from tensorflow.keras.models import Model
import numpy as np
import os
import pickle
from tensorflow.keras.callbacks import EarlyStopping
from tensorflow.keras.models import load_model
from tensorflow.keras.utils import to_categorical
from sklearn.preprocessing import StandardScaler
from sklearn.utils import shuffle
from joblib import dump as j_dump
from joblib import load as j_load
import random
from codes.base import eyeing as ey
from openpyxl import Workbook


class Modeling():
    @staticmethod
    def create_io():
        """
        creating in-out model
        
        Parameters:
            None
        
        Returns:
            None
        """
        
        print("Starting to create an empty in_out model...")
        inp1_shape = (ey.EYE_SIZE[0], ey.EYE_SIZE[1]*2, 1)
        x2_chosen_features = (0, 1, 2, 3, 4, 5, 6, 7, 8, 9)
        inp2_shape = (len(x2_chosen_features),)

        inp1 = Input(inp1_shape)
        layer = Conv2D(16, (11, 11), (1, 1), 'same', activation='relu')(inp1)
        layer = MaxPooling2D((2, 2), (2, 2))(layer)
        layer = Conv2D(32, (7, 7), (1, 1), 'same', activation='relu')(layer)
        layer = MaxPooling2D((2, 2), (2, 2))(layer)
        layer = Conv2D(64, (5, 5), (1, 1), 'same', activation='relu')(layer)
        layer = MaxPooling2D((2, 2), (2, 2))(layer)
        layer = Conv2D(128, (3, 3), (1, 1), activation='relu')(layer)
        layer = MaxPooling2D((2, 2), (2, 2))(layer)
        layer = Flatten()(layer)
        inp2 = Input(inp2_shape)
        layer = Concatenate()([layer, inp2])
        layer = Dense(256, 'relu')(layer)
        layer = Dense(128, 'relu')(layer)
        layer = Dense(32, 'relu')(layer)
        layer = Dense(8, 'relu')(layer)
        output_layer = Dense(1, "sigmoid")(layer)
        input_layers = [inp1, inp2]
        model = Model(inputs=input_layers, outputs=output_layer)
        model.compile(optimizer="adam", loss="binary_crossentropy", metrics="acc")
        print(model.summary())
        n_weights = np.sum([np.prod(v.get_shape()) for v in model.trainable_weights])

        mdl_num = ey.find_max_mdl(ey.io_raw_dir) + 1
        info = {"n_weights": n_weights,
                "input1_shape": inp1_shape,
                "input2_shape": inp2_shape,
                "x2_chosen_features": x2_chosen_features}
        mdl_name = ey.MDL + f"{mdl_num}"
        mdl_dir = ey.io_raw_dir + mdl_name + ".h5"
        model.save(mdl_dir)
        ey.save([info], ey.io_raw_dir, [mdl_name])
        print("\nEmpty in_out model created and saved to " + mdl_dir)


    @staticmethod
    def create_et():
        """
        Creating eye tracking model. You can change the structure in following, as you want.
        
        Parameters:
            None
        
        Returns:
            None
        """
        
        print("Starting to create empty eye_tracking models...")
        inp1_shape = (ey.EYE_SIZE[0], ey.EYE_SIZE[1]*2, 1)
        x2_chosen_features = (0, 1, 2, 3, 4, 5, 6, 7, 8, 9)
        inp2_shape = (len(x2_chosen_features),)

        inp1 = Input(inp1_shape)
        layer = Conv2D(16, (11, 11), (1, 1), 'same', activation='relu')(inp1)
        layer = MaxPooling2D((2, 2), (2, 2))(layer)
        layer = Conv2D(32, (7, 7), (1, 1), 'same', activation='relu')(layer)
        layer = MaxPooling2D((2, 2), (2, 2))(layer)
        layer = Conv2D(64, (5, 5), (1, 1), 'same', activation='relu')(layer)
        layer = MaxPooling2D((2, 2), (2, 2))(layer)
        layer = Conv2D(128, (3, 3), (1, 1), activation='relu')(layer)
        layer = MaxPooling2D((2, 2), (2, 2))(layer)
        layer = Flatten()(layer)
        inp2 = Input(inp2_shape)
        layer = Concatenate()([layer, inp2])
        layer = Dense(256, 'relu')(layer)
        layer = Dense(128, 'relu')(layer)
        layer = Dense(32, 'relu')(layer)
        layer = Dense(8, 'relu')(layer)
        out = Dense(1, 'linear')(layer)
        input_layers = [inp1, inp2]
        model = Model(inputs=input_layers, outputs=out)
        model.compile(optimizer='adam', loss='mse')
        print(model.summary())
        n_weights = np.sum([np.prod(v.get_shape()) for v in model.trainable_weights])

        mdl_num = ey.find_max_mdl(ey.et_raw_dir) + 1
        info = {"n_weights": n_weights,
                "input1_shape": inp1_shape,
                "input2_shape": inp2_shape,
                "x2_chosen_features": x2_chosen_features}

        mdl_name = ey.MDL + f"{mdl_num}"
        mdl_dir = ey.et_raw_dir + mdl_name + ".h5"
        model.save(mdl_dir)
        ey.save([info], ey.et_raw_dir, [mdl_name])
        print("\nEmpty eye_tracking model created and saved to " + mdl_dir)

    @staticmethod
    def train_io(
        subjects,
        models_list,
        min_max_brightness_ratio=[[0.65, 1.45]],
        r_train_list=[0.85],
        n_epochs_patience=[[160, 10]],
        save_scaler=False,
        show_model=False
        ):
        """
        Training the io models. This method uses the dataset in the io folder of subject's number folder. The parameters should be lists.
        So, you can train each model with several parameters and hyper parameters to see which one works better.

        Parameters:
            subjects: a list of subject numbers that you want to train the model with them.
            models_list: You can train several models at a same time. So, you can enter a list of model numbers
            min_max_brightness_ratio: To make the models robust to the brightness, the eyes images are multiplies into a number between two considered numbers
            r_train_list: The ratio for train dataset
            n_epochs_patience: The number of epochs and patience to intrupt training
            save_scaler: To save the scaler
            show_model: To show the model
        
        Returns:
            None
        """
        print("Starting to train in_out model...")
        x1_load = []
        x2_load = []
        y_load = []
        for sbj in subjects:
            data_io_dir = ey.create_dir([ey.subjects_dir, f"{sbj}", ey.IO])
            x1_load0, x2_load0, y_load0 = ey.load(data_io_dir, [ey.X1, ey.X2, ey.Y])
            for (x10, x20, y10) in zip(x1_load0[0], x2_load0[0], y_load0[0]):
                x1_load.append(x10)
                x2_load.append(x20)
                y_load.append(y10)

        x1_load = np.array(x1_load)
        x2_load = np.array(x2_load)
        y_load = np.array(y_load)

        n_smp = x1_load.shape[0]
        print(f"\nNumber of samples : {n_smp}")

        # changing brightness
        j = 1
        for mbr in min_max_brightness_ratio:
            x1_new = x1_load.copy()
            for (i, _) in enumerate(x1_load):
                r = random.uniform(mbr[0], mbr[1])
                x1_new[i] = (x1_new[i] * r).astype(np.uint8)

            for raw_mdl_num in models_list:
                info = ey.load(ey.io_raw_dir, [ey.MDL + f"{raw_mdl_num}"])[0]
                x2_chosen_features = info["x2_chosen_features"]
                x2_new = x2_load[:, x2_chosen_features]

                x1_shf, x2_shf, y_shf = shuffle(x1_new, x2_new, y_load)

                x1_scaler = ey.X1_SCALER
                x1 = x1_shf / x1_scaler

                x2_scaler = StandardScaler()
                x2 = x2_scaler.fit_transform(x2_shf)

                scalers = [x1_scaler, x2_scaler]
                if save_scaler:
                    j_dump(scalers, ey.scalers_dir + f"scl_io_{len(x2_chosen_features)}.bin")

                for rt in r_train_list:
                    n_train = int(rt * n_smp)
                    x1_train, x2_train = x1[:n_train], x2[:n_train]
                    x1_val, x2_val = x1[n_train:], x2[n_train:]
                    
                    y_train = y_shf[:n_train]
                    y_val = y_shf[n_train:]
                    print("\nTrain and val data shape:")
                    print(x1_train.shape, x1_val.shape, x2_train.shape, x2_val.shape,
                          y_train.shape, y_val.shape)

                    x_train = [x1_train, x2_train]
                    x_val = [x1_val, x2_val]

                    for nep in n_epochs_patience:
                        info["min_max_brightness_ratio"] = mbr
                        info["r_train"] = rt
                        info["n_epochs_patience"] = nep
                        cb = EarlyStopping(patience=nep[1], verbose=1, restore_best_weights=True)

                        raw_model_dir = ey.io_raw_dir + ey.MDL + f"{raw_mdl_num}.h5"
                        print("\nLoading blink_in_out model from " + raw_model_dir)
                        model = load_model(raw_model_dir)
                        if show_model:
                            print(model.summary())

                        print(f"\n<<<<<<< {j}-model:{raw_mdl_num}-min_max_ratio:{mbr}-r_train:{rt}-epoch_patience:{nep} >>>>>>>>")
                        model.fit(x_train,
                                  y_train,
                                  validation_data=(x_val, y_val),
                                  epochs=nep[0],
                                  callbacks=cb)
                        train_loss = model.evaluate(x_train, y_train)
                        val_loss = model.evaluate(x_val, y_val)

                        info["train_loss"] = train_loss
                        info["val_loss"] = val_loss

                        trained_mdl_num = ey.find_max_mdl(ey.io_trained_dir) + 1
                        mdl_name = ey.MDL + f'{trained_mdl_num}'
                        ey.save([info], ey.io_trained_dir, [mdl_name])
                        mdl_tr_dir = ey.io_trained_dir + mdl_name + ".h5"
                        model.save(mdl_tr_dir)
                        print("\nSaving in_out model in " + mdl_tr_dir)
                        j += 1
        

    @staticmethod
    def train_et(
        subjects,
        models_list,
        min_max_brightness_ratio=[[0.65, 1.45]],
        r_train_list=[0.8],
        n_epochs_patience=[[100, 15]],
        shift_samples=None,
        blinking_threshold="d",
        save_scaler=False,
        show_model=False
        ):
        """
        Training the et (base) models. This method uses the dataset in the et folder of subject's number folder. The parameters should be lists.
        So, you can train each model with several parameters and hyper parameters to see which one works better.

        Parameters:
            subjects: a list of subject numbers that you want to train the model with them.
            models_list: You can train several models at a same time. So, you can enter a list of model numbers
            min_max_brightness_ratio: To make the models robust to the brightness, the eyes images are multiplies into a number between two considered numbers
            r_train_list: The ratio for train dataset
            n_epochs_patience: The number of epochs and patience to intrupt training
            shift_samples: To shift sample if there is a high delay
            blinking_threshold: It can have three types --> d: default, ao: app offered, uo: user offered
            save_scaler: To save the scaler
            show_model: To show the model
        
        Returns:
            None
        """
        print("Starting to train eye_tracking models...")
        x1_load = []
        x2_load = []
        y_load = []
        kk = 0
        for sbj in subjects:
            sbj_dir = ey.create_dir([ey.subjects_dir, f"{sbj}"])
            sbj_clb_dir = ey.create_dir([sbj_dir, ey.CLB])

            (
                sbj_x1_load,
                sbj_x2_load,
                sbj_y_load,
                sbj_t_mat,
                sbj_eyes_ratio
            ) = ey.load(sbj_clb_dir, [ey.X1, ey.X2, ey.Y, ey.T, ey.ER])

            if shift_samples:
                if shift_samples[kk]:
                    ii = 0
                    for (x11, x21, y1, t1, eyr1) in zip(sbj_x1_load, sbj_x2_load, sbj_y_load, sbj_t_mat, sbj_eyes_ratio):
                        sbj_t_mat[ii] = t1[:-shift_samples[kk]]
                        sbj_x1_load[ii] = x11[shift_samples[kk]:]
                        sbj_x2_load[ii] = x21[shift_samples[kk]:]
                        sbj_y_load[ii] = y1[:-shift_samples[kk]]
                        sbj_eyes_ratio[ii] = eyr1[shift_samples[kk]:]
                        ii += 1

            kk += 1
            sbj_er_dir = ey.create_dir([sbj_dir, ey.ER])
            sbj_blinking_threshold = ey.get_threshold(sbj_er_dir, blinking_threshold)

            sbj_blinking = ey.get_blinking(sbj_t_mat, sbj_eyes_ratio, sbj_blinking_threshold)[1]

            for (x11, x21, y1, b1) in zip(sbj_x1_load, sbj_x2_load, sbj_y_load, sbj_blinking):
                for (x10, x20, y0, b0) in zip(x11, x21, y1, b1):
                    if not b0:
                        x1_load.append(x10)
                        x2_load.append(x20)
                        y_load.append(y0)
        x1_load = np.array(x1_load)
        x2_load = np.array(x2_load)
        y_load = np.array(y_load)
        n_smp = x1_load.shape[0]
        print(f"\nNumber of samples : {n_smp}")
        j = 1
        for mbr in min_max_brightness_ratio:
            x1_new = x1_load.copy()
            for (i, _) in enumerate(x1_load):
                r = random.uniform(mbr[0], mbr[1])
                x1_new[i] = (x1_new[i] * r).astype(np.uint8)

            for raw_mdl_num in models_list:
                info = ey.load(ey.et_raw_dir, [ey.MDL + f"{raw_mdl_num}"])[0]
                x2_chosen_features = info["x2_chosen_features"]
                x2_new = x2_load[:, x2_chosen_features]

                x1_shf, x2_shf, y_hrz_shf, y_vrt_shf = shuffle(x1_new, x2_new, y_load[:, 0], y_load[:, 1])

                x1_scaler = ey.X1_SCALER
                x1 = x1_shf / x1_scaler

                x2_scaler = StandardScaler()
                x2 = x2_scaler.fit_transform(x2_shf)
                y_scaler = ey.Y_SCALER

                scalers = [x1_scaler, x2_scaler, y_scaler]

                if save_scaler:
                    j_dump(scalers, ey.scalers_dir + f"scl_et_{len(x2_chosen_features)}.bin")

                for rt in r_train_list:
                    n_train = int(rt * n_smp)
                    x1_train, x2_train = x1[:n_train], x2[:n_train]
                    x1_val, x2_val = x1[n_train:], x2[n_train:]
                    
                    y_hrz_train, y_vrt_train = y_hrz_shf[:n_train], y_vrt_shf[:n_train]
                    y_hrz_val, y_vrt_val = y_hrz_shf[n_train:], y_vrt_shf[n_train:]
                    print("\nTrain and val data shape:")
                    print(x1_train.shape, x1_val.shape, x2_train.shape, x2_val.shape,
                          y_hrz_train.shape, y_hrz_val.shape, y_vrt_train.shape, y_vrt_val.shape)

                    x_train = [x1_train, x2_train]
                    x_val = [x1_val, x2_val]

                    for nep in n_epochs_patience:
                        info["min_max_brightness_ratio"] = mbr
                        info["r_train"] = rt
                        info["n_epochs_patience"] = nep
                        cb = EarlyStopping(patience=nep[1], verbose=1, restore_best_weights=True)

                        raw_model_dir = ey.et_raw_dir + ey.MDL + f"{raw_mdl_num}.h5"
                        print("\nLoading eye_tracking model from " + raw_model_dir)
                        model_hrz = load_model(raw_model_dir)
                        model_vrt = load_model(raw_model_dir)
                        if show_model:  
                            print(model_hrz.summary())

                        trained_mdl_num = ey.find_max_mdl(ey.et_trained_dir, b=-7) + 1

                        print(f"\n<<<<<<< {j}-model-hrz:{raw_mdl_num}-min_max_ratio:{mbr}-r_train:{rt}-epoch_patience:{nep} >>>>>>>>")
                        model_hrz.fit(x_train,
                                      y_hrz_train * y_scaler,
                                      validation_data=(x_val, y_hrz_val * y_scaler),
                                      epochs=nep[0],
                                      callbacks=cb)
                        mdl_name = ey.MDL + f"{trained_mdl_num}"
                        mdl_hrz_tr_dir = ey.et_trained_dir + mdl_name + "-hrz.h5"
                        print("\nSaving horizontally eye_tracking model in " + mdl_hrz_tr_dir)
                        model_hrz.save(mdl_hrz_tr_dir)
                        hrz_train_loss = model_hrz.evaluate(x_train, y_hrz_train * y_scaler)
                        hrz_val_loss = model_hrz.evaluate(x_val, y_hrz_val * y_scaler)
                        info["hrz_train_loss"] = hrz_train_loss
                        info["hrz_val_loss"] = hrz_val_loss

                        print(f"\n<<<<<<< {j}-model-vrt:{raw_mdl_num}-min_max_ratio:{mbr}-r_train:{rt}-epoch_patience:{nep} >>>>>>>>")
                        model_vrt.fit(x_train,
                                      y_vrt_train * y_scaler,
                                      validation_data=(x_val, y_vrt_val * y_scaler),
                                      epochs=nep[0],
                                      callbacks=cb)
                        tr_model_vrt_dir = ey.et_trained_dir + mdl_name + f"-vrt.h5"
                        print("Saving vertically eye_tracking model in " + tr_model_vrt_dir)
                        model_vrt.save(tr_model_vrt_dir)
                        vrt_train_loss = model_vrt.evaluate(x_train, y_vrt_train * y_scaler)
                        vrt_val_loss = model_vrt.evaluate(x_val, y_vrt_val * y_scaler)
                        info["vrt_train_loss"] = vrt_train_loss
                        info["vrt_val_loss"] = vrt_val_loss

                        ey.save([info], ey.et_trained_dir, [mdl_name])

                        j += 1


    @staticmethod
    def get_models_information(io=True, raw=True, show_model=False):
        """
        To write the models information in an excel file. It gets the information from attached pickle file for each model

        Parameters:
            io: If it's io or et
            raw: If the model is trained or not
            show_model: If you want to show the model

        Returns:
            None
        """
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "# of model"
        ws['B1'] = "# of weights"
        ws['C1'] = "input 1 shape"
        ws['D1'] = "input 2 shape"
        ws['E1'] = "x2 chosen features"
        if io:
            if raw:
                files_name = os.listdir(ey.io_raw_dir)
                if files_name:
                    for fn in files_name:
                        if fn[-7:] == ".pickle":
                            mdl_num = int(fn[3:-7])
                            mdl_name = ey.MDL + f"{mdl_num}"
                            if show_model:
                                mdl_dir = ey.io_raw_dir + mdl_name + ".h5"
                                mdl = load_model(mdl_dir)
                                print(f"<<<<<<<<<<<<<< {mdl_dir} >>>>>>>>>>>>>>")
                                print(mdl.summary())
                            info = ey.load(ey.io_raw_dir, [mdl_name])[0]

                            ws[f'A{mdl_num+1}'] = str(mdl_num)
                            ws[f'B{mdl_num+1}'] = str(info['n_weights'])
                            ws[f'C{mdl_num+1}'] = str(info['input1_shape'])
                            ws[f'D{mdl_num+1}'] = str(info['input2_shape'])
                            ws[f'E{mdl_num+1}'] = str(info['x2_chosen_features'])
            else:
                ws['F1'] = "min-Max brightness ratio"
                ws['G1'] = "r_train"
                ws['H1'] = "# of epochs and patience"
                ws['I1'] = "train loss"
                ws['J1'] = "val loss"

                files_name = os.listdir(ey.io_trained_dir)
                if files_name:
                    for fn in files_name:
                        if fn[-7:] == ".pickle":
                            mdl_num = int(fn[3:-7])
                            mdl_name = ey.MDL + f"{mdl_num}"
                            if show_model:
                                mdl_dir = ey.io_trained_dir + mdl_name + ".h5"
                                mdl = load_model(mdl_dir)
                                print(f"<<<<<<<<<<<<<< {mdl_dir} >>>>>>>>>>>>>>")
                                print(mdl.summary())
                            info = ey.load(ey.io_trained_dir, [mdl_name])[0]

                            ws[f'A{mdl_num+1}'] = str(mdl_num)
                            ws[f'B{mdl_num+1}'] = str(info['n_weights'])
                            ws[f'C{mdl_num+1}'] = str(info['input1_shape'])
                            ws[f'D{mdl_num+1}'] = str(info['input2_shape'])
                            ws[f'E{mdl_num+1}'] = str(info['x2_chosen_features'])
                            ws[f'F{mdl_num+1}'] = str(info['min_max_brightness_ratio'])
                            ws[f'G{mdl_num+1}'] = str(info['r_train'])
                            ws[f'H{mdl_num+1}'] = str(info['n_epochs_patience'])
                            ws[f'I{mdl_num+1}'] = str(info['train_loss'])
                            ws[f'J{mdl_num+1}'] = str(info['val_loss'])

        else:
            if raw:
                files_name = os.listdir(ey.et_raw_dir)
                if files_name:
                    for fn in files_name:
                        if fn[-7:] == ".pickle":
                            mdl_num = int(fn[3:-7])
                            mdl_name = ey.MDL + f"{mdl_num}"
                            if show_model:
                                mdl_dir = ey.et_raw_dir + mdl_name + ".h5"
                                mdl = load_model(mdl_dir)
                                print(f"<<<<<<<<<<<<<< {mdl_dir} >>>>>>>>>>>>>>")
                                print(mdl.summary())
                            info = ey.load(ey.et_raw_dir, [mdl_name])[0]

                            ws[f'A{mdl_num+1}'] = str(mdl_num)
                            ws[f'B{mdl_num+1}'] = str(info['n_weights'])
                            ws[f'C{mdl_num+1}'] = str(info['input1_shape'])
                            ws[f'D{mdl_num+1}'] = str(info['input2_shape'])
                            ws[f'E{mdl_num+1}'] = str(info['x2_chosen_features'])

            else:
                ws['F1'] = "min-Max brightness ratio"
                ws['G1'] = "r_train"
                ws['H1'] = "# of epochs and patience"
                ws['I1'] = "model-hrz train loss"
                ws['J1'] = "model-hrz val loss"
                ws['K1'] = "model-vrt train loss"
                ws['L1'] = "model-vrt val loss"

                files_name = os.listdir(ey.et_trained_dir)
                if files_name:
                    for fn in files_name:
                        if fn[-7:] == ".pickle":
                            mdl_num = int(fn[3:-7])
                            mdl_name = ey.MDL + f"{mdl_num}"
                            if show_model:
                                mdl_dir = ey.et_trained_dir + mdl_name + "-hrz.h5"
                                mdl = load_model(mdl_dir)
                                print(f"<<<<<<<<<<<<<< {mdl_dir} >>>>>>>>>>>>>>")
                                print(mdl.summary())
                            info = ey.load(ey.et_trained_dir, [mdl_name])[0]

                            ws[f'A{mdl_num+1}'] = str(mdl_num)
                            ws[f'B{mdl_num+1}'] = str(info['n_weights'])
                            ws[f'C{mdl_num+1}'] = str(info['input1_shape'])
                            ws[f'D{mdl_num+1}'] = str(info['input2_shape'])
                            ws[f'E{mdl_num+1}'] = str(info['x2_chosen_features'])
                            ws[f'F{mdl_num+1}'] = str(info['min_max_brightness_ratio'])
                            ws[f'G{mdl_num+1}'] = str(info['r_train'])
                            ws[f'H{mdl_num+1}'] = str(info['n_epochs_patience'])
                            ws[f'I{mdl_num+1}'] = str(info['hrz_train_loss'])
                            ws[f'J{mdl_num+1}'] = str(info['hrz_val_loss'])
                            ws[f'K{mdl_num+1}'] = str(info['vrt_train_loss'])
                            ws[f'L{mdl_num+1}'] = str(info['vrt_val_loss'])

        if io and raw:
            info_name = "info_io_raw"
        elif io and not raw:
            info_name = "info_io_trained"
        elif not io and raw:
            info_name = "info_et_raw"
        else:
            info_name = "info_et_trained"

        wb.save(ey.files_dir + info_name + ".xlsx")