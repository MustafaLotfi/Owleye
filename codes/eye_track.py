import pickle
from tensorflow.keras.models import load_model
import numpy as np
from joblib import load as j_load
from codes.base import eyeing as ey
from scipy import signal
from openpyxl import Workbook, load_workbook
import os


class EyeTrack(object):
    @staticmethod
    def get_pixels(
        subjects,
        models_list=[1],
        target_fol=ey.SMP,
        shift_samples=None,
        blinking_threshold="uo",
        use_io=False,
        delete_files=False
        ):
        tfn = 1
        if target_fol == ey.ACC:
            tfn = 2
        elif target_fol == ey.LTN:
            tfn = 3

        out_threshold_min = 0.005
        out_threshold_max = 0.995
        latency_radius = 0.33
        median_filter_window_size = 5

        x1_scaler_et, x2_scaler_et, y_scaler = j_load(ey.scalers_dir + "scalers_et_main.bin")
        if tfn == 1:
            x1_scaler_io, x2_scaler_io = j_load(ey.scalers_dir + "scalers_io_main.bin")
            mdl_io = load_model(ey.io_trained_dir + ey.MDL + "1.h5")

        kk = 0
        for num in subjects:
            print(f"<<<<<<<<<<<<<<<<<<<<< Subject {num} >>>>>>>>>>>>>>>>>>>>>>>")
            sbj_dir = ey.create_dir([ey.subjects_dir, f"{num}"])
            sbj_models_dir = ey.create_dir([sbj_dir, ey.MDL])
            target_dir = ey.create_dir([sbj_dir, target_fol])
            if ey.file_existing(target_dir, ey.X1+".pickle"):
                if tfn == 1:
                    t_load, sys_time_load, x1_load, x2_load, eyes_ratio = ey.load(target_dir, [ey.T, "sys_time", ey.X1, ey.X2, ey.ER])
                    if shift_samples:
                        if shift_samples[kk]:
                            ii = 0
                            for (x11, x21, t1, st1, eyr1) in zip(x1_load, x2_load, t_load, sys_time_load, eyes_ratio):
                                t_load[ii] = t1[:-shift_samples[kk]]
                                sys_time_load[ii] = st1[:-shift_samples[kk]]
                                x1_load[ii] = x11[shift_samples[kk]:]
                                x2_load[ii] = x21[shift_samples[kk]:]
                                eyes_ratio[ii] = eyr1[shift_samples[kk]:]
                                ii += 1
                elif tfn == 2:
                    t_load, x1_load, x2_load, y_load, eyes_ratio = ey.load(target_dir, [ey.T, ey.X1, ey.X2, ey.Y, ey.ER])
                    if shift_samples:
                        if shift_samples[kk]:
                            ii = 0
                            for (x11, x21, y1, t1, eyr1) in zip(x1_load, x2_load, y_load, t_load, eyes_ratio):
                                t_load[ii] = t1[:-shift_samples[kk]]
                                x1_load[ii] = x11[shift_samples[kk]:]
                                x2_load[ii] = x21[shift_samples[kk]:]
                                y_load[ii] = y1[:-shift_samples[kk]]
                                eyes_ratio[ii] = eyr1[shift_samples[kk]:]
                                ii += 1
                else:
                    t_load, x1_load, x2_load = ey.load(target_dir, [ey.T, ey.X1, ey.X2])
                    if shift_samples:
                        if shift_samples[kk]:
                            ii = 0
                            for (x11, x21, t1) in zip(x1_load, x2_load, t_load):
                                t_load[ii] = t1[:-shift_samples[kk]]
                                x1_load[ii] = x11[shift_samples[kk]:]
                                x2_load[ii] = x21[shift_samples[kk]:]
                                ii += 1
                kk += 1
                for mdl_num in models_list:
                    mdl_et_name = ey.MDL + f"{mdl_num}"
                    mdl_et_hrz_dir = sbj_models_dir + mdl_et_name + "-hrz.h5"
                    mdl_et_vrt_dir = sbj_models_dir + mdl_et_name + "-vrt.h5"
                    if ey.file_existing(sbj_models_dir, mdl_et_name + "-hrz.h5"):
                        info = ey.load(sbj_models_dir, [mdl_et_name])[0]
                        x2_chosen_features = info["x2_chosen_features"]
                        mdl_et_hrz = load_model(mdl_et_hrz_dir)
                        mdl_et_vrt = load_model(mdl_et_vrt_dir)

                        y_prd = []
                        for (x11, x21) in zip(x1_load, x2_load):
                            n_smp_vec = x11.shape[0]
                            x21_new = x21[:, x2_chosen_features]
                            x11_nrm = x11 / x1_scaler_et
                            x21_nrm = x2_scaler_et.transform(x21_new)
                            x0_nrm = [x11_nrm, x21_nrm]

                            y_hrz_prd = np.expand_dims(mdl_et_hrz.predict(x0_nrm).reshape((n_smp_vec,)), 1) / y_scaler
                            y_vrt_prd = np.expand_dims(mdl_et_vrt.predict(x0_nrm).reshape((n_smp_vec,)), 1) / y_scaler

                            y_prd.append(np.concatenate([y_hrz_prd, y_vrt_prd], 1))

                        if tfn == 3:
                            t_delay = []
                            j = 0
                            for (t1, y1_prd) in zip(t_load, y_prd):
                                for (t0, y0_prd) in zip(t1, y1_prd):
                                    if j % 2 == 0:
                                        d = y0_prd[0] - 0.66
                                    else:
                                        d = 0.33 - y0_prd[0]
                                    if 0 < d < latency_radius:
                                        t_delay.append(t0 - t1[0])
                                        break
                                j += 1
                            print(t_delay)
                            t_delay = np.array(t_delay).mean() - ey.LATENCY_WAITING_TIME/1000.0
                            print(t_delay)
                            ey.save([t_delay], target_dir, ["t_delay"])

                        else:
                            # get out data
                            y_in = y_prd.copy()
                            if (tfn == 1) and use_io:
                                for (x11, x21, yi1) in zip(x1_load, x2_load, y_in):
                                    x1_io = x11 / x1_scaler_io
                                    x2_io = x2_scaler_io.transform(x21)
                                    y_io_prd = mdl_io.predict([x1_io, x2_io]).round()
                                    
                                    for (et0, yio) in zip(yi1, y_io_prd):
                                        if yio == 1:
                                            et0[0] = -1
                                            et0[1] = -1

                            er_dir = ey.create_dir([sbj_dir, ey.ER])
                            blinking_threshold = ey.get_threshold(er_dir, blinking_threshold)
                            blinking = ey.get_blinking(t_load, eyes_ratio, blinking_threshold)[1]
                            for (yi1, bl1) in zip(y_in, blinking):
                                for (yi0, bl0) in zip(yi1, bl1):
                                    if bl0:
                                        yi0[0] = -1
                                        yi0[1] = -1

                            y_prd_mat = []
                            for yi1 in y_in:
                                blinking_out = (yi1[:, 0] == -1)
                                n_smp = yi1.shape[0]
                                i = 0
                                while i < (n_smp):
                                    bo_vec = []
                                    in_vec = []
                                    now = blinking_out[i]
                                    if now:
                                        bo_vec.append(yi1[i])
                                    else:
                                        in_vec.append(yi1[i])
                                    j = 1
                                    if (i+j) < n_smp:
                                        while blinking_out[i+j] == now:
                                            if now:
                                                bo_vec.append(yi1[i+j])
                                            else:
                                                in_vec.append(yi1[i+j])
                                            j += 1
                                            if (i+j) >= n_smp:
                                                break
                                    if now:
                                        y_prd_mat.append(np.array(bo_vec))
                                    else:
                                        y_prd_mat.append(np.array(in_vec))
                                    i += j

                            for y_prd_vec in y_prd_mat:
                                if y_prd_vec[0, 0] != -1:
                                    if 3 < y_prd_vec.shape[0] < (median_filter_window_size+2):
                                        y_prd_vec[:, 0] = signal.medfilt(y_prd_vec[:, 0], 3)
                                        y_prd_vec[:, 1] = signal.medfilt(y_prd_vec[:, 1], 3)
                                    elif y_prd_vec.shape[0] >= (median_filter_window_size+2):
                                        y_prd_vec[:, 0] = signal.medfilt(y_prd_vec[:, 0], median_filter_window_size)
                                        y_prd_vec[:, 1] = signal.medfilt(y_prd_vec[:, 1], median_filter_window_size)

                            # # Concatenating y
                            y_prd_fnl = y_prd_mat[0]
                            for (i, y_prd_vec) in enumerate(y_prd_mat):
                                if i == 0:
                                    continue
                                y_prd_fnl = np.concatenate([y_prd_fnl, y_prd_vec], 0)

                            if tfn == 1:
                                t = []
                                sys_time = []
                                for (t1, st1) in zip(t_load, sys_time_load):
                                    for (t0, st0) in zip(t1, st1):
                                        t.append(t0)
                                        sys_time.append(st0)
                                t = np.array(t)
                                wb = Workbook()
                                ws = wb.active
                                ws['A1'] = "Time"
                                ws['A2'] = "sec"
                                ws['B1'] = "SystemTime"
                                ws['C1'] = "EyeTrack"
                                ws['C2'] = "(p_x/scr_w,p_y/scr_h)"
                                ws['D1'] = "Condition"
                                ws['D2'] = "{start,stop}"
                                ws['D3'] = "start"
                                for i in range(y_prd_fnl.shape[0]):
                                    ws[f'A{i+3}'] = f"{t[i]}"
                                    ws[f'B{i+3}'] = sys_time[i]
                                    ws[f'C{i+3}'] = f"({round(y_prd_fnl[i, 0] * 10000)/10000},{round(y_prd_fnl[i, 1] * 10000)/10000})"
                                ws[f'D{i+3}'] = "stop"
                                wb.save(target_dir + "eye_track.xlsx")
                                ey.save([t, y_prd_fnl], target_dir, ["t_vec", "y_prd"])

                                if delete_files:
                                    ey.remove(target_dir, [ey.FV])
                            else:
                                y_vec = []
                                for y1 in y_load:
                                    for y0 in y1:
                                        y_vec.append(y0)
                                y_vec = np.array(y_vec)

                                y_vec = y_vec[y_prd_fnl[:, 0] != -1]
                                y_prd_fnl = y_prd_fnl[y_prd_fnl[:, 0] != -1]

                                losses = np.sum(((y_prd_fnl-y_vec)*y_scaler)**2, 0) / y_vec.shape[0]

                                print(f"Lossess for two hrz and vrt models: {losses}")

                                info["hrz_retrain_test_loss"] = losses[0]
                                info["vrt_retrain_test_loss"] = losses[1]

                                y_prd_fnl[y_prd_fnl < out_threshold_min] = out_threshold_min
                                y_prd_fnl[y_prd_fnl > out_threshold_max] = out_threshold_max

                                ey.save([info], sbj_models_dir, [mdl_et_name])
                                ey.save([y_vec, y_prd_fnl], target_dir, ["y_mdf", "y_prd_mdf"])

                                if delete_files:
                                    ey.remove(target_dir, [ey.Y])
                            
                            if delete_files:
                                ey.remove(sbj_models_dir)
                                ey.remove(target_dir, [ey.ER, ey.X1, ey.X2, ey.T])
                    else:
                        print(f"Data does not exist in {sbj_models_dir}")
            else:
                print(f"Data does not exist in {target_dir}")


    @staticmethod
    def get_fixations(
            subjects,
            n_monitors_data=1,
            t_discard=0.1,
            x_merge=0.2/2,
            y_merge=0.25/2,
            vx_thr=2.5,
            vy_thr=2.5
    ):
          # pxr  --> pixel ratio (pixel/screen_width)
          # pxr
          # pxr/sec
          # pxr/sec

        for num in subjects:
            smp_dir = ey.create_dir([ey.subjects_dir, f"{num}", ey.SMP])

            if ey.file_existing(smp_dir, "eye_track.xlsx"):
                sheet = load_workbook(smp_dir + "eye_track.xlsx")["Sheet"]
                max_row = sheet.max_row
                et_xl = []
                for i in range(3, max_row+1):
                    et_cell_list = sheet[f"C{i}"].value[1:-1].split(',')
                    et_xl.append(
                        [float(sheet[f"A{i}"].value),
                         sheet[f"B{i}"].value,
                         float(et_cell_list[0]),
                         float(et_cell_list[1]),
                         sheet[f"D{i}"].value]
                        )
                n_smp_all = len(et_xl)

                i = 0
                t_mat_seq = []
                t_sys_mat_seq = []
                et_mat_seq = []
                while i < n_smp_all:
                    if (et_xl[i][4] == "start") or (et_xl[i][4] == "Start"):
                        t1 = []
                        ts1 = []
                        et1 = []
                        j = 0
                        while True:
                            t1.append([et_xl[i+j][0]])
                            ts1.append([et_xl[i+j][1]])
                            et1.append([et_xl[i+j][2], et_xl[i+j][3]])
                            if et_xl[i+j][4] == "stop" or et_xl[i+j][4] == "Stop":
                                break
                            j += 1
                        t_mat_seq.append(np.array(t1).reshape((len(t1),)))
                        t_sys_mat_seq.append(ts1)
                        et_mat_seq.append(np.array(et1))
                        i += j
                    i += 1

                t = t_mat_seq[0]
                t_sys = t_sys_mat_seq[0]
                et = et_mat_seq[0]
                for (i, t1) in enumerate(t_mat_seq):
                    if i == 0:
                        continue
                    t = np.concatenate([t, t1])
                    t_sys += t_sys_mat_seq[i]
                    et = np.concatenate([et, et_mat_seq[i]])

                t_mat = []
                t_sys_mat = []
                et_mat = []
                for (t1, ts1, et1) in zip(t_mat_seq, t_sys_mat_seq, et_mat_seq):
                    n_smp1 = t1.shape[0]
                    blinking_out = (et1[:, 0] == -1)
                    t_mat1 = []
                    ts_mat1 = []
                    et_mat1 = []
                    i = 0
                    while i < (n_smp1):
                        t0 = [t1[i]]
                        ts0 = [ts1[i]]
                        bo_vec = []
                        in_vec = []
                        now = blinking_out[i]
                        if now:
                            bo_vec.append(et1[i])
                        else:
                            in_vec.append(et1[i])
                        j = 1
                        if (i+j) < n_smp1:
                            while blinking_out[i+j] == now:
                                t0.append(t1[i+j])
                                ts0.append(ts1[i+j])
                                if now:
                                    bo_vec.append(et1[i+j])
                                else:
                                    in_vec.append(et1[i+j])
                                j += 1
                                if (i+j) >= n_smp1:
                                    break
                        t_mat1.append(np.array(t0))
                        ts_mat1.append(ts0)
                        if now:
                            et_mat1.append(np.array(bo_vec))
                        else:
                            et_mat1.append(np.array(in_vec))
                        i += j
                    t_mat.append(t_mat1)
                    t_sys_mat.append(ts_mat1)
                    et_mat.append(et_mat1)

                saccades = []
                vet_mat = []
                for (t2, et2) in zip(t_mat, et_mat):
                    saccades1 = []
                    vet_mat1 = []
                    for (t1, et1) in zip(t2, et2):
                        if et1[0, 0] != -1:
                            if et1.shape[0] == 1:
                                vet1 = np.zeros((1,2))
                                s1 = [None]
                            else:
                                vet1 = et1.copy()
                                vet1[1:, 0] = (et1[1:, 0] - et1[:-1, 0]) / (t1[1:] - t1[:-1])
                                vet1[1:, 1] = (et1[1:, 1] - et1[:-1, 1]) / (t1[1:] - t1[:-1])
                                vet1[0] = vet1[1]

                                s1 = ((vet1[:, 0]>vx_thr)+(vet1[:, 0]<-vx_thr))+((vet1[:, 1]>vy_thr)+(vet1[:, 1]<-vy_thr))
                        else:
                            et_shape = et1.shape[0]
                            vet1 = np.zeros(et1.shape)
                            s1 = np.array([None] * et_shape)
                        vet_mat1.append(vet1)
                        saccades1.append(s1)
                    saccades.append(saccades1)
                    vet_mat.append(vet_mat1)


                vet4 = []
                for vet3 in vet_mat:
                    vet2 = vet3[0].copy()
                    for (i, vet1) in enumerate(vet3):
                        if i == 0:
                            continue
                        vet2 = np.concatenate([vet2, vet1], 0)
                    vet4.append(np.array(vet2))

                vet = vet4[0]
                for (i, vet1) in enumerate(vet4):
                    if i == 0:
                        continue
                    vet = np.concatenate([vet, vet1])


                sac_mat_new = []
                t_mat_new = []
                t_sys_mat_new = []
                et_mat_new = []
                for (t_mat1, ts_mat1, et_mat1, saccades1) in zip(t_mat, t_sys_mat, et_mat, saccades):
                    k = 0
                    sac_mat_new1 = []
                    t_mat_new1 = []
                    t_sys_mat_new1 = []
                    et_mat_new1 = []
                    for (t1, ts1, et1, sac1) in zip(t_mat1, ts_mat1, et_mat1, saccades1):
                        if et1[0, 0] != -1:
                            n_smp = t1.shape[0]
                            i = 0
                            while i < (n_smp):
                                s0 = [sac1[i]]
                                t0 = [t1[i]]
                                ts0 = [ts1[i]]
                                et0 = [et1[i]]
                                now = sac1[i]
                                j = 1
                                if (i+j) < n_smp:
                                    while sac1[i+j] == now:
                                        s0.append(sac1[i+j])
                                        t0.append(t1[i+j])
                                        ts0.append(ts1[i+j])
                                        et0.append(et1[i+j])
                                        j += 1
                                        if (i+j) >= n_smp:
                                            break
                                sac_mat_new1.append(np.array(s0))
                                t_mat_new1.append(np.array(t0))
                                t_sys_mat_new1.append(ts0)
                                et_mat_new1.append(np.array(et0))
                                i += j
                        else:
                            sac_mat_new1.append(sac1)
                            t_mat_new1.append(t1)
                            t_sys_mat_new1.append(ts1)
                            et_mat_new1.append(et1)
                    sac_mat_new.append(sac_mat_new1)
                    t_mat_new.append(t_mat_new1)
                    t_sys_mat_new.append(t_sys_mat_new1)
                    et_mat_new.append(et_mat_new1)


                fix = []
                k = 0
                for (sac_mat_new1, t_mat_new1, t_sys_mat_new1, et_mat_new1) in zip(sac_mat_new, t_mat_new, t_sys_mat_new, et_mat_new):
                    fix1 = []
                    for (s1, t1, ts1, et1) in zip(sac_mat_new1, t_mat_new1, t_sys_mat_new1, et_mat_new1):
                        sac_shp = s1.shape
                        if s1[0] == False:
                            if not s1[0]:
                                fix1.append([k,
                                             sac_shp[0],
                                             t1[0],
                                             round(t1[-1]-t1[0], 2),
                                             round(et1[:, 0].mean(), 4),
                                             round(et1[:, 1].mean(), 4),
                                             ts1[0]])
                        k += sac_shp[0]
                    fix.append(fix1)

                fix_mrg_one = []
                for fix1 in fix:
                    fix_mrg1 = []
                    n_fix = len(fix1)
                    i = 0
                    while i < n_fix:
                        f_new = fix1[i]
                        j = 1
                        while (i+j) < n_fix:
                            fj = fix1[i+j]
                            fj_d = ((fj[4]-f_new[4])/(x_merge/n_monitors_data))**2+((fj[5]-f_new[5])/(y_merge))**2
                            if fj_d < 1:
                                f_new = [f_new[0],
                                         f_new[1] + fj[1],
                                         f_new[2],
                                         round(f_new[3] + fj[3], 2),
                                         round((f_new[4]*f_new[1]+fj[4]*fj[1])/(f_new[1]+fj[1]), 4),
                                         round((f_new[5]*f_new[1]+fj[5]*fj[1])/(f_new[1]+fj[1]), 4),
                                         f_new[-1]]
                                if (i+j) == n_fix-1:
                                    fix_mrg1.append(f_new)
                                not_joined = False
                            else:
                                fix_mrg1.append(f_new)
                                not_joined = True
                                break
                            j += 1
                        i += j
                    if not_joined:
                        fix_mrg1.append(fix1[-1])
                        
                    fix_mrg_one.append(fix_mrg1)


                fix_dcd = []
                for fix_mrg1 in fix_mrg_one:
                    fix_dcd1 = []
                    for f in fix_mrg1:
                        if f[3] >= t_discard:
                            fix_dcd1.append(f)
                    fix_dcd.append(fix_dcd1)


                fix_mrg_two = []
                for fix1 in fix_dcd:
                    fix_mrg1 = []
                    n_fix = len(fix1)
                    i = 0
                    while i < n_fix:
                        f_new = fix1[i]
                        j = 1
                        while (i+j) < n_fix:
                            fj = fix1[i+j]
                            fj_d = ((fj[4]-f_new[4])/(x_merge/n_monitors_data))**2+((fj[5]-f_new[5])/(y_merge))**2
                            if fj_d < 1:
                                f_new = [f_new[0],
                                         f_new[1] + fj[1],
                                         f_new[2],
                                         round(f_new[3] + fj[3], 2),
                                         round((f_new[4]*f_new[1]+fj[4]*fj[1])/(f_new[1]+fj[1]), 4),
                                         round((f_new[5]*f_new[1]+fj[5]*fj[1])/(f_new[1]+fj[1]), 4),
                                         f_new[-1]]
                                if (i+j) == n_fix-1:
                                    fix_mrg1.append(f_new)
                                not_joined = False
                            else:
                                fix_mrg1.append(f_new)
                                not_joined = True
                                break
                            j += 1
                #             if (i+j) >= n_fix:
                #                 break
                        i += j
                    if not_joined:
                        fix_mrg1.append(fix1[-1])
                        
                    fix_mrg_two.append(fix_mrg1)

                wb = Workbook()
                ws = wb.active
                ws['A1'] = "FixationTime"
                ws['A2'] = "sec"
                ws['B1'] = "FixationSystemTime"
                ws['C1'] = "FixationDuration"
                ws['C2'] = "sec"
                ws['D1'] = "FixationLocation"
                ws['D2'] = "(p_x/scr_w,p_y/scr_h)"
                i = 0
                for f_seq in fix_mrg_two:
                    for f in f_seq:
                        ws[f'A{i+3}'] = f"{f[2]}"
                        ws[f'B{i+3}'] = f[6][0]
                        ws[f'C{i+3}'] = f"{f[3]}"
                        ws[f'D{i+3}'] = f"({f[4]},{f[5]})"
                        i += 1
                
                wb.save(smp_dir + "fixations.xlsx")
            else:
                print(f"Data does not exist in {smp_dir}")

    @staticmethod
    def get_models_information(show_model=False):
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "subject"
        ws['B1'] = "model"
        ws['C1'] = "trained model"
        ws['D1'] = "weights"
        ws['E1'] = "input 1 shape"
        ws['F1'] = "input 2 shape"
        ws['G1'] = "x2 chosen features"
        ws['H1'] = "min-Max brightness ratio"
        ws['I1'] = "r_train"
        ws['J1'] = "epochs and patience"
        ws['K1'] = "model-hrz train loss"
        ws['L1'] = "model-hrz val loss"
        ws['M1'] = "model-vrt train loss"
        ws['N1'] = "model-vrt val loss"
        ws['O1'] = "r_retrain"
        ws['P1'] = "epochs and patience-retrain"
        ws['Q1'] = "trainable layers"
        ws['R1'] = "model-hrz-retrain train loss"
        ws['S1'] = "model-hrz-retrain val loss"
        ws['T1'] = "model-vrt-retrain train loss"
        ws['U1'] = "model-vrt-retrain val loss"
        ws['V1'] = "model-hrz-retrain test loss"
        ws['W1'] = "model-vrt-retrain test loss"


        subjects = os.listdir(ey.subjects_dir)
        i = 2
        for sbj in subjects:
            sbj = int(sbj)
            models_dir = ey.create_dir([ey.subjects_dir, f"{sbj}", ey.MDL])
            files_name = os.listdir(models_dir)
            if files_name:
                for fn in files_name:
                    if fn[-7:] == ".pickle":
                        mdl_num = int(fn[3:-7])
                        mdl_name = ey.MDL + f"{mdl_num}"
                        if show_model:
                            mdl = load_model(models_dir + mdl_name + "-hrz.h5")
                            print(mdl.summary())
                        info = ey.load(models_dir, [ey.MDL + f"{mdl_num}"])[0]

                        ws[f'A{i}'] = str(sbj)
                        ws[f'B{i}'] = str(mdl_num)
                        ws[f'C{i}'] = str(info['trained_mdl_num'])
                        ws[f'D{i}'] = str(info['n_weights'])
                        ws[f'E{i}'] = str(info['input1_shape'])
                        ws[f'F{i}'] = str(info['input2_shape'])
                        ws[f'G{i}'] = str(info['x2_chosen_features'])
                        ws[f'H{i}'] = str(info['min_max_brightness_ratio'])
                        ws[f'I{i}'] = str(info['r_train'])
                        ws[f'J{i}'] = str(info['n_epochs_patience'])
                        ws[f'K{i}'] = str(info['hrz_train_loss'])
                        ws[f'L{i}'] = str(info['hrz_val_loss'])
                        ws[f'M{i}'] = str(info['vrt_train_loss'])
                        ws[f'N{i}'] = str(info['vrt_val_loss'])
                        ws[f'O{i}'] = str(info['r_retrain'])
                        ws[f'P{i}'] = str(info['n_epochs_patience_retrain'])
                        ws[f'Q{i}'] = str(info['trainable_layers'])
                        ws[f'R{i}'] = str(info['hrz_retrain_train_loss'])
                        ws[f'S{i}'] = str(info['hrz_retrain_val_loss'])
                        ws[f'T{i}'] = str(info['vrt_retrain_train_loss'])
                        ws[f'U{i}'] = str(info['vrt_retrain_val_loss'])
                        ws[f'V{i}'] = str(info['hrz_retrain_test_loss'])
                        ws[f'W{i}'] = str(info['vrt_retrain_test_loss'])

                        i += 1

        wb.save(ey.files_dir + "info_et_retrains.xlsx")